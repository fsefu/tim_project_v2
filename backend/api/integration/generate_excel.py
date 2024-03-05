import json
from django.http import JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def generate_excel(lists_data, segments_data):
    wb = Workbook()
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)  # Merge cells A1:C1
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Segements and List Data"
    title_cell.font = Font(size=16, bold=True)  # Change font size and make it bold
    title_cell.fill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')  # Change background color
    title_cell.alignment = Alignment(horizontal='center')  # Center align the title

    # Add headers
    headers = ["ID", "Name", "Created At", "Updated At", "Members", "Type", "Definition Link"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)  # Make the header row bold
        cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # Change header row background color
  

    # Adding lists data
    for i in range(len(lists_data['ID'])):
        ws.append([lists_data['ID'][i], lists_data['Name'][i], lists_data['Created At'][i], lists_data['Updated At'][i], lists_data['Members'][i], lists_data['Type'][i], lists_data['Definition Link'][i]])

    # Adding segments data
    for i in range(len(segments_data['ID'])):
        ws.append([segments_data['ID'][i], segments_data['Name'][i], segments_data['Created At'][i], segments_data['Updated At'][i], segments_data['Members'][i], segments_data['Type'][i], segments_data['Definition Link'][i]])

    # Auto-size columns based on content
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column if cell.value]
        if column:
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Adding a small buffer and scaling factor for better visibility
            ws.column_dimensions[cell.column_letter].width = adjusted_width

    # Add hyperlink style to the "Definition Link" column
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            if cell.value:
                cell.hyperlink = cell.value

    combined_data = {}
    for key in lists_data.keys():
        combined_data[key] = lists_data[key] + segments_data[key]

    # Write the combined data into a new JSON file
    with open('api/data/lists_segments_data.json', 'w') as combined_file:
        json.dump(combined_data, combined_file, indent=4)

    wb.save("api/data/lists_and_segments.xlsx")

    return JsonResponse({'detail': 'Successfully Updated.'})


